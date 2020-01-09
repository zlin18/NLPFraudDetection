# -*- coding: utf-8 -*-
"""
Created on Mon Mar 19 20:49:46 2018

@author: Duo Cao
"""

import jieba
import jieba.posseg as pseg
from optparse import OptionParser
from openpyxl import load_workbook, Workbook

flag = True  # True 为真货， False 为假货

file_name = "comment_test.xlsx"
wb = load_workbook(filename=file_name)
sheet = wb.worksheets[0]

M = sheet.max_row
N = sheet.max_column

nip_library = ['x','uj','ul','r']
dict_id = 0
dict_list = []
word_library = []
for col in sheet.iter_cols(min_row=1,min_col=2,max_row=M,max_col=N):
    temp_library = {}
    for cell in col: 
        temp_str = cell.value
        if temp_str == None:
            continue
       # print(temp_str)
        seg_list = pseg.cut(temp_str)
        
        for word, flag in seg_list:
            if flag not in nip_library:
                if word in temp_library:
                    temp_library[word] += 1
                else:
                    temp_library[word] = 1
    for key in temp_library:
        if key not in word_library and temp_library[key]>1:
            #print(key,temp_library[key])
            word_library.append(key)
            
    dict_list.append(temp_library)
    dict_id += 1

#creating a new workbook
new_wb = Workbook()
new_ws = new_wb.active
new_ws.append(["id"]+word_library+["TF"])

if flag:
    index = 1
    for dict in dict_list:
        new_row = [str(index)]
        for word in word_library:
            if word in dict:
                new_row.append(dict[word])
            else:
                new_row.append(0)
        index +=1
        new_row.append("True")
        new_ws.append(new_row)
else:
    index = 1
    for dict in dict_list:
        new_row = [str(index)]
        for word in word_library:
            if word in dict:
                new_row.append(dict[word])
            else:
                new_row.append(0)
        index +=1
        new_row.append("False")
        new_ws.append(new_row)
new_wb.save("build_frequency.xlsx")